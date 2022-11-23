import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from pathlib import Path


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        correct_price = cell.value * 0.9
        correct_price_cell = sheet.cell(row, 4)
        correct_price_cell.value = correct_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    # Change bar filling and line color
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "00000"
    s.graphicalProperties.solidFill = "ff9900"
    sheet.add_chart(chart, 'e2')
    wb.save(filename)
    # let's add a comment to see changes in git


path = Path()
for file in path.glob('*.xlsx'):
    print(file)
    process_workbook(file)
